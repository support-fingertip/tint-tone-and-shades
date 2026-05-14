"""Generate the Odoo 19 bank transaction import template (.xlsx).

Run:  python3 build_template.py
Output: odoo19_bank_transaction_import_template.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "odoo19_bank_transaction_import_template.xlsx"

# Real importable fields on account.bank.statement.line in Odoo 19.
# - partner_id            : Many2one to res.partner; importing with the partner's
#                           display name lets Odoo's import wizard auto-match.
# - partner_bank_id/acc_number : sub-field path so the IBAN/account string
#                           resolves the Many2one to res.partner.bank.
# - journal_id            : Many2one to account.journal; supply the bank
#                           journal name so the line lands on the right journal.
# Removed from the previous template:
# - partner_name          : not a field (caused "Set Partner" on every row).
# - transaction_type      : not a field; Odoo derives sign from `amount`.
HEADERS = [
    "date",
    "journal_id",
    "payment_ref",
    "amount",
    "partner_id",
    "partner_bank_id/acc_number",
]

SAMPLES = [
    ("2025-01-03", "Bank", "NEFT-UTR-001234567890", 12500.00, "Infosys Ltd", "IN60SBIN0000123456789"),
    ("2025-01-05", "Bank", "UPI-9876543210@okaxis", -3200.50, "Amazon India", ""),
    ("2025-01-08", "Bank", "NACH-BESCOM-20250108", -1500.00, "BESCOM", ""),
    ("2025-01-10", "Bank", "NEFT-UTR-001234567891", 5000.00, "Tata Consultancy", "IN60HDFC0001234567890"),
    ("2025-01-12", "Bank", "UPI-8765432109@paytm", -800.00, "Swiggy", ""),
    ("2025-01-15", "Bank", "CHQ-000123", -9500.00, "HDFC Bank", ""),
    ("2025-01-20", "Bank", "RTGS-UTR-AB2025012001", 25000.00, "Wipro Ltd", "IN60ICIC0009876543210"),
    ("2025-01-25", "Bank", "UPI-7654321098@gpay", -2100.75, "Zepto", ""),
]

INSTRUCTIONS = [
    ("Odoo 19 — Bank Statement Import Template", ""),
    ("All 6 columns map directly to fields on account.bank.statement.line. Data starts at Row 2 — never delete Row 1.", ""),
    ("", ""),
    ("IMPORTANT — use the RIGHT import path", "Do NOT use 'Accounting → Bank journal → Import'. That path runs the bank-statement file parser (CSV/OFX/CAMT/QIF) which only auto-links partners via IBAN — it ignores partner_id and leaves every line with 'Set Partner' / 'Set Account' buttons. Use Odoo's generic data import on the account.bank.statement.line model instead — see STEP 4."),
    ("", ""),
    ("STEP 1", "Delete sample rows 2–9 from the 'Bank Transactions' sheet. Keep Row 1 exactly as-is — Odoo uses it for column auto-mapping."),
    ("STEP 2", "Paste your real transactions from Row 2 downward.\nDate: YYYY-MM-DD  |  Amount: positive = credit (money in), negative = debit (money out)  |  payment_ref must be unique per row.\nEvery partner_id value must match the display name of an existing res.partner in Odoo (case-sensitive). Create the partner first if it doesn't exist."),
    ("STEP 3", "Save the file as Excel Workbook (.xlsx). Do not rename the sheet or change the field names in Row 1."),
    ("STEP 4", "In Odoo 19, enable Developer Mode (Settings → Developer Tools → Activate developer mode).\nOpen this URL in the same browser tab — it loads the bank statement line list view:\n    /odoo/action-account.action_bank_statement_line\nClick the ⚙ (cog) / Actions menu → 'Import records' → upload this .xlsx → keep sheet 'Bank Transactions' and 'Use first row as headers' on.\nThis path uses Odoo's base_import which DOES resolve Many2one fields like partner_id and journal_id by display name."),
    ("STEP 5", "Click 'Test'. Every row should show resolved partner_id and journal_id (no 'partner Infosys Ltd not found' style warnings). Then click 'Import'.\nOpen the Bank Matching widget — imported lines will already show partner + 'Reconcile' / 'Receivable', exactly like a manually-created line."),
    ("", ""),
    ("Field Reference", ""),
    ("date", "Transaction date. ISO format YYYY-MM-DD. Example: 2025-03-31."),
    ("journal_id", "Bank journal display name (e.g. 'Bank'). base_import resolves it by name_search; must match an existing account.journal of type=bank."),
    ("payment_ref", "Unique bank reference: UTR / NEFT / IMPS / UPI Ref ID / cheque number. Shown as the label on the statement line."),
    ("amount", "Signed amount. Positive (+) = money in (credit). Negative (−) = money out (debit)."),
    ("partner_id", "Customer or vendor display name. base_import resolves it to res.partner by name_search. Setting the partner lets Odoo derive the counterpart account automatically — that's what removes 'Set Partner' and 'Set Account' from the Bank Matching widget."),
    ("partner_bank_id/acc_number", "Counterparty IBAN / bank account number. Sub-field path lets the Many2one to res.partner.bank resolve. Optional — leave blank if unknown."),
    ("", ""),
    ("Why the previous template kept showing 'Set Partner' / 'Set Account'", ""),
    ("Wrong import path", "The 'Bank journal → Import' button runs the bank-statement file parser, not base_import. That parser only auto-links partners via IBAN, and ignores partner_id even if you supply it. For UPI/UTR-only counterparties (Zepto, Swiggy, BESCOM…) it can't resolve a partner at all → 'Set Partner' button on every row."),
    ("partner_name (old column)", "partner_name IS a real Char field on account.bank.statement.line, but it's only the raw counterparty text from the statement — it does not link to res.partner. Replaced with partner_id, which base_import resolves to a real partner record."),
    ("transaction_type (old column)", "Not a field on account.bank.statement.line. Odoo derives credit vs debit from the sign of 'amount'. Removed."),
    ("partner_bank_id (raw IBAN, old column)", "partner_bank_id is a Many2one — importing a raw IBAN string under that bare column name doesn't resolve. Use partner_bank_id/acc_number so base_import maps to the res.partner.bank record."),
]


def build():
    wb = Workbook()

    # --- Sheet 1: Bank Transactions ---
    ws = wb.active
    ws.title = "Bank Transactions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5BBA", end_color="2E5BBA", fill_type="solid")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r, row in enumerate(SAMPLES, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = border
            if HEADERS[c - 1] == "amount":
                cell.number_format = "#,##0.00;[Red]-#,##0.00"

    widths = {
        "date": 12,
        "journal_id": 14,
        "payment_ref": 28,
        "amount": 14,
        "partner_id": 22,
        "partner_bank_id/acc_number": 28,
    }
    for col_idx, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 18)
    ws.freeze_panes = "A2"

    # --- Sheet 2: Instructions ---
    info = wb.create_sheet("Instructions")
    title_font = Font(bold=True, size=14, color="2E5BBA")
    label_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    info.column_dimensions["A"].width = 32
    info.column_dimensions["B"].width = 100

    for r, (left, right) in enumerate(INSTRUCTIONS, start=1):
        a = info.cell(row=r, column=1, value=left)
        b = info.cell(row=r, column=2, value=right)
        a.alignment = wrap
        b.alignment = wrap
        if right == "" and left and r == 1:
            a.font = title_font
        elif right == "" and left:
            a.font = label_font
        else:
            a.font = label_font

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    build()
